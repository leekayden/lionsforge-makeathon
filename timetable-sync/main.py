import inquirer
import xlrd
import sys
import os
import keyboard
import time
import webbrowser
from termcolor import colored
import subprocess
import pandas as pd
from openpyxl import Workbook

print(r"""
 _____ _                _        _     _        __                  _                     _              
/__   (_)_ __ ___   ___| |_ __ _| |__ | | ___  / _\_   _ _ __   ___| |__  _ __ ___  _ __ (_)_______ _ __ 
  / /\/ | '_ ` _ \ / _ \ __/ _` | '_ \| |/ _ \ \ \| | | | '_ \ / __| '_ \| '__/ _ \| '_ \| |_  / _ \ '__|
 / /  | | | | | | |  __/ || (_| | |_) | |  __/ _\ \ |_| | | | | (__| | | | | | (_) | | | | |/ /  __/ |   
 \/   |_|_| |_| |_|\___|\__\__,_|_.__/|_|\___| \__/\__, |_| |_|\___|_| |_|_|  \___/|_| |_|_/___\___|_|   
                                                   |___/                                                 
""")
time.sleep(.5)
print()
time.sleep(.5)
print(colored('[i] ', 'blue') + "Welcome to the Timetable Synchronizer")
time.sleep(1)
print(colored('[c] ', 'blue') + "(c) Copyright 2022 Kayden Lee (Cloudserve Tech)")
time.sleep(1)
print(colored('[v] ', 'blue') + "v1.2 Pre-release Dev Build 1762022")
time.sleep(1)
print(colored('[www] ', 'blue') + "https://cloudservetechcentral.com/")
time.sleep(3)
print()
print(r"""
Steps:
===========================================================
1. Download sample timetable
2. Upload timetable
    > Use default settings
    > Use custom settings
3. Check values
4. Parse values
5. Send values to raspberry pi
6. Download values for debugging or future use (optional)
""")

time.sleep(2)

print(colored('[?] ', 'yellow') + "Press any key to continue...")
keyboard.read_key()
time.sleep(3)
print()


def one():
    print(colored('[1] ', 'magenta') + "Download sample timetable\n")
    print(colored('[i] ', 'blue') + "Detecting browser...")
    time.sleep(.5)
    print(colored('[i] ', 'blue') + "Downloading sample timetable...")
    time.sleep(2)
    # webbrowser.open('https://cdn.cloudservetechcentral.com/timetable-sync/sample-timetable.xslx')
    # file = input(colored('[!] ', 'red') + "Sample timetable has not been uploaded to the server yet!")
    print(colored('[!] ', 'red') + "Sample timetable has not been uploaded to the server yet!")


def two():
    print(colored('[2] ', 'magenta') + "Upload timetable\n")
    time.sleep(2)
    file_name = input(colored('[i] ', 'blue') + "Please enter the exact location of the timetable\n" + colored('[?] ', 'yellow') + ">_ ")
    time.sleep(2)
    print(colored('[i] ', 'blue') + "Opening file in explorer... Please ensure that the path of your excel document is correct.")
    time.sleep(2)
    subprocess.Popen(r'explorer /select,' + file_name)

    print()

    df = pd.read_excel(file_name, index_col=0)
    print(df.head())

    print("\n" + colored('[i] ', 'blue') + "This is the excel file that you have selected. Is this correct?")
    response = None
    while response not in {"yes", "no"}:
        response = input(colored('[?] ', 'yellow') + "Please enter yes or no: ").lower()

    if response == "yes":
        three()
    elif response == "no":
        print(colored('[i] ', 'blue') + "Restarting step 2...")
        time.sleep(2.5)
        two()
    else:
        print(colored('[!] ', 'There was a critical error with the program! Please install the latest LTS version from https://software.cloudservetechcentral.com/download/timetable-sync'))

def three():
    print()
    time.sleep(2)
    wb = Workbook()
    ws = wb.active
    print(colored('[3] ', 'magenta') + "Check values")
    print(colored('[i] ', 'blue') + "Checking values...")
    for cell in ws['A']:
        print(cell.value)
    cell_range = ws['A2':'A5']

one()
print()
time.sleep(2)
two()
